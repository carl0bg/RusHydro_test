from pathlib import Path

from pandas.core.frame import DataFrame 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def apply_styles_to_excel(file_path: str, column_name: str):
    """Добавление цвета в зависимости от значения в столбце"""
    wb = load_workbook(file_path)
    ws = wb.active

    column_index = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == column_name:
            column_index = idx
            break

    if column_index is None:
        raise ValueError(f"Столбец '{column_name}' не найден в файле Excel.")

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=column_index, max_col=column_index):
        for cell in row:
            if cell.value == 0:
                cell.fill = green_fill
            else:
                cell.fill = red_fill

    wb.save(file_path)



def add_new_columns(data_file: DataFrame) -> DataFrame:
    '''Выполнение логики с добавлением новых столбцов'''
    data_file['Налог Исчислено по формуле'] = data_file['Налоговая база'].apply(
            lambda x: x * 0.13 if x <= 5000000 else x * 0.15
        )
    data_file['Отклонения'] = data_file['Налог'] - data_file['Налог Исчислено по формуле']
    sorted_data = data_file.sort_values(by='Отклонения', ascending=False)
    return sorted_data



def update_existing_template(data_file: DataFrame):
    '''Добавление данных к файлу-шаблону'''
    template_path = Path("template_file/rept_header.xlsx")
    wb = load_workbook(template_path)
    ws = wb.active

    for index, row in data_file.iterrows():
        ws.append(row.values.tolist())

    updated_file_path = Path("static/new.xlsx")
    wb.save(updated_file_path)
    apply_styles_to_excel("static/new.xlsx", 'Отклонения')